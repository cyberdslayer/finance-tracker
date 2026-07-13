from rest_framework import generics, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from django.http import HttpResponse, FileResponse
from django.db.models import Sum, Count
from datetime import datetime
import csv
import io
from io import BytesIO
from transactions.models import Transaction
from budgets.models import Budget
from .models import Report
from .serializers import ReportSerializer


def generate_pdf_report(user, month, year):
    """Generate PDF report using reportlab"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from datetime import datetime as dt

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1976d2'),
        spaceAfter=30,
        alignment=1
    )
    story.append(Paragraph(f"Financial Report - {dt(year, month, 1).strftime('%B %Y')}", title_style))
    story.append(Spacer(1, 0.2*inch))

    # Summary Section
    income = Transaction.objects.filter(
        user=user, type='income', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    expenses = Transaction.objects.filter(
        user=user, type='expense', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    story.append(Paragraph("Summary", styles['Heading2']))
    summary_data = [
        ['Metric', 'Amount'],
        ['Total Income', f'₹ {float(income):.2f}'],
        ['Total Expenses', f'₹ {float(expenses):.2f}'],
        ['Net Savings', f'₹ {float(income - expenses):.2f}'],
    ]

    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))

    # Category Breakdown
    story.append(Paragraph("Category-wise Expense Breakdown", styles['Heading2']))
    category_data = [['Category', 'Amount', 'Transactions']]
    
    categories = Transaction.objects.filter(
        user=user, type='expense', date__month=month, date__year=year
    ).values('category').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')

    for cat in categories:
        category_data.append([
            cat['category'].replace('_', ' ').title(),
            f'₹ {float(cat["total"]):.2f}',
            str(cat['count'])
        ])

    category_table = Table(category_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
    category_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(category_table)
    story.append(Spacer(1, 0.3*inch))

    # Budget Status
    story.append(PageBreak())
    story.append(Paragraph("Budget Status", styles['Heading2']))
    
    budgets = Budget.objects.filter(user=user, month=month, year=year)
    if budgets.exists():
        budget_data = [['Category', 'Limit', 'Spent', 'Status']]
        for budget in budgets:
            status = 'Over Budget' if budget.is_over_budget else f'{budget.percentage_used:.1f}% Used'
            budget_data.append([
                budget.get_category_display(),
                f'₹ {float(budget.monthly_limit):.2f}',
                f'₹ {float(budget.spent):.2f}',
                status
            ])
        
        budget_table = Table(budget_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        budget_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(budget_table)

    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph(f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_csv_report(user, month, year):
    """Generate CSV report"""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Header
    writer.writerow(['Personal Finance Report', month, year])
    writer.writerow([])

    # Summary
    income = Transaction.objects.filter(
        user=user, type='income', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    expenses = Transaction.objects.filter(
        user=user, type='expense', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    writer.writerow(['Summary'])
    writer.writerow(['Total Income', income])
    writer.writerow(['Total Expenses', expenses])
    writer.writerow(['Net Savings', income - expenses])
    writer.writerow([])

    # Transactions
    writer.writerow(['Transactions'])
    writer.writerow(['Date', 'Type', 'Category', 'Amount', 'Description'])
    
    transactions = Transaction.objects.filter(
        user=user, date__month=month, date__year=year
    ).order_by('-date')
    
    for trans in transactions:
        writer.writerow([
            trans.date,
            trans.get_type_display(),
            trans.get_category_display(),
            trans.amount,
            trans.description or ''
        ])

    buffer.seek(0)
    return buffer


def generate_excel_report(user, month, year):
    """Generate Excel report"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    workbook = openpyxl.Workbook()
    
    # Summary Sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    summary_sheet['A1'] = f"Financial Report - {datetime(year, month, 1).strftime('%B %Y')}"
    summary_sheet['A1'].font = Font(bold=True, size=14, color="1976D2")
    
    summary_sheet['A3'] = "Metric"
    summary_sheet['B3'] = "Amount"
    
    for cell in ['A3', 'B3']:
        summary_sheet[cell].fill = header_fill
        summary_sheet[cell].font = header_font
    
    income = Transaction.objects.filter(
        user=user, type='income', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    expenses = Transaction.objects.filter(
        user=user, type='expense', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    summary_sheet['A4'] = "Total Income"
    summary_sheet['B4'] = float(income)
    summary_sheet['A5'] = "Total Expenses"
    summary_sheet['B5'] = float(expenses)
    summary_sheet['A6'] = "Net Savings"
    summary_sheet['B6'] = float(income - expenses)
    
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 15

    # Transactions Sheet
    trans_sheet = workbook.create_sheet("Transactions")
    trans_sheet['A1'] = "Date"
    trans_sheet['B1'] = "Type"
    trans_sheet['C1'] = "Category"
    trans_sheet['D1'] = "Amount"
    trans_sheet['E1'] = "Description"
    
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
        trans_sheet[cell].fill = header_fill
        trans_sheet[cell].font = header_font
    
    transactions = Transaction.objects.filter(
        user=user, date__month=month, date__year=year
    ).order_by('-date')
    
    row = 2
    for trans in transactions:
        trans_sheet[f'A{row}'] = trans.date
        trans_sheet[f'B{row}'] = trans.get_type_display()
        trans_sheet[f'C{row}'] = trans.get_category_display()
        trans_sheet[f'D{row}'] = float(trans.amount)
        trans_sheet[f'E{row}'] = trans.description or ''
        row += 1
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        trans_sheet.column_dimensions[col].width = 18

    # Budget Sheet
    budget_sheet = workbook.create_sheet("Budgets")
    budget_sheet['A1'] = "Category"
    budget_sheet['B1'] = "Monthly Limit"
    budget_sheet['C1'] = "Spent"
    budget_sheet['D1'] = "Remaining"
    budget_sheet['E1'] = "Status"
    
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
        budget_sheet[cell].fill = header_fill
        budget_sheet[cell].font = header_font
    
    budgets = Budget.objects.filter(user=user, month=month, year=year)
    row = 2
    for budget in budgets:
        budget_sheet[f'A{row}'] = budget.get_category_display()
        budget_sheet[f'B{row}'] = float(budget.monthly_limit)
        budget_sheet[f'C{row}'] = float(budget.spent)
        budget_sheet[f'D{row}'] = float(budget.remaining)
        budget_sheet[f'E{row}'] = 'Over Budget' if budget.is_over_budget else 'On Track'
        row += 1
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        budget_sheet.column_dimensions[col].width = 18

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def monthly_report(request):
    """Get monthly financial report"""
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    if not month or not year:
        today = datetime.now()
        month = today.month
        year = today.year

    month, year = int(month), int(year)
    user = request.user

    income = Transaction.objects.filter(
        user=user, type='income', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    expenses = Transaction.objects.filter(
        user=user, type='expense', date__month=month, date__year=year
    ).aggregate(total=Sum('amount'))['total'] or 0

    return Response({
        'month': month,
        'year': year,
        'total_income': float(income),
        'total_expenses': float(expenses),
        'net_savings': float(income - expenses),
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_pdf(request):
    """Export monthly report as PDF"""
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    if not month or not year:
        today = datetime.now()
        month = today.month
        year = today.year

    month, year = int(month), int(year)

    pdf_buffer = generate_pdf_report(request.user, month, year)
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="finance_report_{year}_{month:02d}.pdf"'
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_csv(request):
    """Export monthly report as CSV"""
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    if not month or not year:
        today = datetime.now()
        month = today.month
        year = today.year

    month, year = int(month), int(year)

    csv_buffer = generate_csv_report(request.user, month, year)
    response = HttpResponse(csv_buffer.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="finance_report_{year}_{month:02d}.csv"'
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_excel(request):
    """Export monthly report as Excel"""
    month = request.query_params.get('month')
    year = request.query_params.get('year')

    if not month or not year:
        today = datetime.now()
        month = today.month
        year = today.year

    month, year = int(month), int(year)

    excel_buffer = generate_excel_report(request.user, month, year)
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="finance_report_{year}_{month:02d}.xlsx"'
    return response
